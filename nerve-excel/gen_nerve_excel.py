import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Color definitions
NAVY_FILL = PatternFill('solid', fgColor='1B3A5C')
BLUE_FILL = PatternFill('solid', fgColor='2C5AA0')
LIGHT_BLUE_FILL = PatternFill('solid', fgColor='E8F0FE')
WHITE_FONT = Font(name='游ゴシック', bold=True, color='FFFFFF', size=11)
HEADER_FONT = Font(name='游ゴシック', bold=True, color='1B3A5C', size=11)
NORMAL_FONT = Font(name='游ゴシック', size=11)
BOLD_FONT = Font(name='游ゴシック', bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

W_FILLS = {
    0: PatternFill('solid', fgColor='FFFFFF'),
    1: PatternFill('solid', fgColor='DCEEFB'),
    2: PatternFill('solid', fgColor='B0D4F1'),
    3: PatternFill('solid', fgColor='6BAED6'),
    4: PatternFill('solid', fgColor='2171B5'),
}
W_FONTS = {
    0: Font(name='游ゴシック', size=11, color='999999'),
    1: Font(name='游ゴシック', size=11, color='333333'),
    2: Font(name='游ゴシック', size=11, color='333333'),
    3: Font(name='游ゴシック', size=11, color='FFFFFF', bold=True),
    4: Font(name='游ゴシック', size=11, color='FFFFFF', bold=True),
}

UPPER_DATA = {
    'levels': ['C5','C6','C7','C8','T1'],
    'muscles': [
        ('前鋸筋', '長胸神経', [2,1,1,0,0]),
        ('棘下筋', '肩甲上神経', [4,3,0,0,0]),
        ('広背筋', '胸背神経', [0,1,4,1,0]),
        ('三角筋', '腋窩神経', [4,1,0,0,0]),
        ('上腕二頭筋', '筋皮神経', [4,3,0,0,0]),
        ('上腕三頭筋', '橈骨神経', [0,1,4,3,0]),
        ('腕橈骨筋', '橈骨神経', [4,1,0,0,0]),
        ('総指伸筋', '橈骨神経', [0,0,3,4,0]),
        ('尺側手根伸筋', '橈骨神経', [0,1,4,2,0]),
        ('長母指伸筋', '橈骨神経', [0,0,1,4,1]),
        ('短母指伸筋', '橈骨神経', [0,0,1,4,1]),
        ('円回内筋', '正中神経', [0,4,1,0,0]),
        ('浅指屈筋', '正中神経', [0,0,1,3,4]),
        ('深指屈筋(示)', '正中神経', [0,0,0,1,4]),
        ('長母指屈筋', '正中神経', [0,0,0,3,4]),
        ('方形回内筋', '正中神経', [0,0,0,1,4]),
        ('短母指外転筋', '正中神経', [0,0,0,1,4]),
        ('深指屈筋(小)', '尺骨神経', [0,0,0,4,2]),
        ('小指外転筋', '尺骨神経', [0,0,0,4,4]),
        ('橈側手根伸筋', '橈骨神経', [0,4,0,0,0]),
        ('尺側手根屈筋', '尺骨神経', [0,0,4,3,1]),
        ('第一背側骨間筋', '尺骨神経', [0,0,0,4,2]),
        ('第一掌側骨間筋', '尺骨神経', [0,0,0,3,2]),
    ]
}

LOWER_DATA = {
    'levels': ['L1','L2','L3','L4','L5','S1','S2'],
    'muscles': [
        ('腸腰筋', '大腿神経/腰神経叢', [1,1,1,0,0,0,0]),
        ('中殿筋', '上殿神経', [0,0,0,1,4,1,0]),
        ('大殿筋', '下殿神経', [0,0,0,0,1,2,1]),
        ('前脛骨筋', '深腓骨神経', [0,0,0,2,4,0,0]),
        ('後脛骨筋', '脛骨神経', [0,0,0,1,4,0,0]),
        ('長趾屈筋', '脛骨神経', [0,0,0,0,4,2,0]),
        ('内転筋群', '閉鎖神経', [0,2,2,2,0,0,0]),
        ('大腿四頭筋', '大腿神経', [0,0,3,3,0,0,0]),
        ('大腿屈筋群', '脛骨神経/総腓骨神経', [0,0,0,0,2,2,1]),
        ('足趾背屈', '深腓骨神経', [0,0,0,0,4,0,0]),
        ('母趾背屈', '深腓骨神経', [0,0,0,0,4,0,0]),
        ('腓骨筋群', '浅腓骨神経', [0,0,0,0,4,1,0]),
        ('下腿三頭筋', '脛骨神経', [0,0,0,0,2,4,0]),
        ('短趾屈筋', '脛骨神経', [0,0,0,0,0,4,2]),
        ('長母趾屈筋', '脛骨神経', [0,0,0,0,4,2,0]),
        ('短母趾屈筋', '脛骨神経', [0,0,0,0,0,4,2]),
    ]
}

def style_cell(cell, font=None, fill=None, alignment=None, border=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border

def create_workbook(data, title_prefix, output_path):
    wb = openpyxl.Workbook()

    # === Sheet 1: 入力 ===
    ws_input = wb.active
    ws_input.title = '入力'
    levels = data['levels']
    muscles = data['muscles']
    num_levels = len(levels)
    num_muscles = len(muscles)

    # Title row
    ws_input.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3+num_levels)
    title_cell = ws_input.cell(row=1, column=1, value=f'{title_prefix} 神経障害部位推測ツール（修正版）')
    style_cell(title_cell, Font(name='游ゴシック', bold=True, color='FFFFFF', size=14), NAVY_FILL,
               Alignment(horizontal='center', vertical='center'))
    ws_input.row_dimensions[1].height = 36
    for c in range(2, 4+num_levels):
        style_cell(ws_input.cell(row=1, column=c), fill=NAVY_FILL)

    # Header row
    headers = ['筋肉名', '支配神経', 'MMT(0-5)'] + levels
    for i, h in enumerate(headers):
        cell = ws_input.cell(row=2, column=i+1, value=h)
        style_cell(cell, WHITE_FONT, BLUE_FILL, Alignment(horizontal='center', vertical='center'), THIN_BORDER)
    ws_input.row_dimensions[2].height = 28

    # Data rows
    for r, (muscle, nerve, weights) in enumerate(muscles):
        row = r + 3
        # Muscle name
        cell = ws_input.cell(row=row, column=1, value=muscle)
        style_cell(cell, BOLD_FONT, LIGHT_BLUE_FILL, Alignment(vertical='center'), THIN_BORDER)
        # Nerve
        cell = ws_input.cell(row=row, column=2, value=nerve)
        style_cell(cell, NORMAL_FONT, None, Alignment(vertical='center'), THIN_BORDER)
        # MMT input (default 5 = normal)
        cell = ws_input.cell(row=row, column=3, value=5)
        style_cell(cell, Font(name='游ゴシック', size=12, bold=True, color='0000FF'),
                   PatternFill('solid', fgColor='FFFFF0'),
                   Alignment(horizontal='center', vertical='center'), THIN_BORDER)
        # Weights
        for w_i, w in enumerate(weights):
            cell = ws_input.cell(row=row, column=4+w_i, value=w)
            style_cell(cell, W_FONTS[w], W_FILLS[w], Alignment(horizontal='center', vertical='center'), THIN_BORDER)

    # Column widths
    ws_input.column_dimensions['A'].width = 18
    ws_input.column_dimensions['B'].width = 22
    ws_input.column_dimensions['C'].width = 12
    for i in range(num_levels):
        ws_input.column_dimensions[get_column_letter(4+i)].width = 8

    # === Sheet 2: 結果 ===
    ws_result = wb.create_sheet('結果')

    # Title
    ws_result.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(num_levels, 4))
    title_cell = ws_result.cell(row=1, column=1, value='解析結果')
    style_cell(title_cell, Font(name='游ゴシック', bold=True, color='FFFFFF', size=14), NAVY_FILL,
               Alignment(horizontal='center', vertical='center'))
    ws_result.row_dimensions[1].height = 36
    for c in range(2, max(num_levels, 4)+1):
        style_cell(ws_result.cell(row=1, column=c), fill=NAVY_FILL)

    # Section 1: Spinal Level Scores
    ws_result.cell(row=3, column=1, value='■ 脊髄レベル別スコア')
    style_cell(ws_result.cell(row=3, column=1), HEADER_FONT)

    # Level headers
    for i, lv in enumerate(levels):
        cell = ws_result.cell(row=4, column=i+1, value=lv)
        style_cell(cell, WHITE_FONT, BLUE_FILL, Alignment(horizontal='center'), THIN_BORDER)

    # Score formulas: SUMPRODUCT((5-MMT)*weight) for each level
    mmt_col = 'C'  # column C in 入力 sheet
    for i in range(num_levels):
        weight_col = get_column_letter(4+i)
        first_row = 3
        last_row = 2 + num_muscles
        formula = f"=SUMPRODUCT((5-入力!{mmt_col}{first_row}:{mmt_col}{last_row})*入力!{weight_col}{first_row}:{weight_col}{last_row})"
        cell = ws_result.cell(row=5, column=i+1, value=formula)
        style_cell(cell, Font(name='游ゴシック', size=14, bold=True, color='1B3A5C'),
                   LIGHT_BLUE_FILL, Alignment(horizontal='center'), THIN_BORDER)

    # Section 2: Nerve Impairment
    ws_result.cell(row=7, column=1, value='■ 末梢神経別 障害筋数')
    style_cell(ws_result.cell(row=7, column=1), HEADER_FONT)

    # Get unique nerves
    nerves = list(dict.fromkeys(m[1] for m in muscles))

    ws_result.cell(row=8, column=1, value='神経名')
    style_cell(ws_result.cell(row=8, column=1), WHITE_FONT, BLUE_FILL, Alignment(horizontal='center'), THIN_BORDER)
    ws_result.cell(row=8, column=2, value='障害筋数')
    style_cell(ws_result.cell(row=8, column=2), WHITE_FONT, BLUE_FILL, Alignment(horizontal='center'), THIN_BORDER)
    ws_result.cell(row=8, column=3, value='支配筋数')
    style_cell(ws_result.cell(row=8, column=3), WHITE_FONT, BLUE_FILL, Alignment(horizontal='center'), THIN_BORDER)

    for n_i, nerve in enumerate(nerves):
        row = 9 + n_i
        ws_result.cell(row=row, column=1, value=nerve)
        style_cell(ws_result.cell(row=row, column=1), BOLD_FONT, None, None, THIN_BORDER)

        nerve_col = 'B'
        mmt_col = 'C'
        first_row = 3
        last_row = 2 + num_muscles

        # Count impaired: nerve matches AND MMT < 5
        formula_impaired = f'=COUNTIFS(入力!{nerve_col}{first_row}:{nerve_col}{last_row},"{nerve}",入力!{mmt_col}{first_row}:{mmt_col}{last_row},"<5")'
        cell = ws_result.cell(row=row, column=2, value=formula_impaired)
        style_cell(cell, Font(name='游ゴシック', size=12, bold=True, color='DC3545'),
                   None, Alignment(horizontal='center'), THIN_BORDER)

        # Count total muscles for this nerve
        formula_total = f'=COUNTIF(入力!{nerve_col}{first_row}:{nerve_col}{last_row},"{nerve}")'
        cell = ws_result.cell(row=row, column=3, value=formula_total)
        style_cell(cell, NORMAL_FONT, None, Alignment(horizontal='center'), THIN_BORDER)

    # Column widths for result sheet
    ws_result.column_dimensions['A'].width = 24
    ws_result.column_dimensions['B'].width = 14
    ws_result.column_dimensions['C'].width = 14

    # Instruction note
    note_row = 9 + len(nerves) + 2
    ws_result.cell(row=note_row, column=1, value='※ 入力シートのMMT値を変更すると自動で結果が更新されます')
    style_cell(ws_result.cell(row=note_row, column=1), Font(name='游ゴシック', size=10, color='888888'))

    wb.save(output_path)
    print(f'Created: {output_path}')

# Generate both files
base = r'C:\Users\jsber\OneDrive\Documents\Claude_task'
create_workbook(UPPER_DATA, '上肢', f'{base}\\上肢脊髄LV神経推測ツール_修正版.xlsx')
create_workbook(LOWER_DATA, '下肢', f'{base}\\下肢脊髄LV神経推測ツール_修正版.xlsx')
print('Done!')
