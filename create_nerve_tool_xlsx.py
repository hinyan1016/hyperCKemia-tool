from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ===== Color constants =====
NAVY = '1B3A5C'
BLUE = '2C5AA0'
WHITE = 'FFFFFF'
LIGHT_YELLOW = 'FFFFF0'
LIGHT_BLUE_1 = 'E8F4FD'
LIGHT_BLUE_2 = 'C8E0F8'
LIGHT_BLUE_3 = '93C5F8'
GRAY = 'CCCCCC'
RED = 'DC3545'
GREEN = '10B981'
PURPLE = '8B5CF6'
ORANGE = 'E8850C'

navy_fill = PatternFill('solid', fgColor=NAVY)
blue_fill = PatternFill('solid', fgColor=BLUE)
input_fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
white_font = Font(name='Arial', color=WHITE, bold=True)
header_font = Font(name='Arial', color=WHITE, bold=True, size=10)
title_font = Font(name='Arial', color=WHITE, bold=True, size=16)
subtitle_font = Font(name='Arial', color=WHITE, size=10)
section_font = Font(name='Arial', bold=True, size=12, color=NAVY)
label_font = Font(name='Arial', bold=True, size=10)
normal_font = Font(name='Arial', size=10)
input_font = Font(name='Arial', size=12, bold=True, color=RED)
bar_blue = Font(name='Arial', color=BLUE, size=10)
bar_orange = Font(name='Arial', color=ORANGE, size=10)
bar_green = Font(name='Arial', color=GREEN, size=10)
bar_purple = Font(name='Arial', color=PURPLE, size=10)

thin_border = Border(
    left=Side(style='thin', color='D0D8E4'),
    right=Side(style='thin', color='D0D8E4'),
    top=Side(style='thin', color='D0D8E4'),
    bottom=Side(style='thin', color='D0D8E4')
)
red_border = Border(
    left=Side(style='dashed', color=RED),
    right=Side(style='dashed', color=RED),
    top=Side(style='dashed', color=RED),
    bottom=Side(style='dashed', color=RED)
)

weight_fills = {
    0: (PatternFill('solid', fgColor='F5F5F5'), Font(name='Arial', color=GRAY, size=9)),
    1: (PatternFill('solid', fgColor=LIGHT_BLUE_1), Font(name='Arial', color=BLUE, bold=True, size=9)),
    2: (PatternFill('solid', fgColor=LIGHT_BLUE_2), Font(name='Arial', color=NAVY, bold=True, size=9)),
    3: (PatternFill('solid', fgColor=LIGHT_BLUE_3), Font(name='Arial', color='0D2B4D', bold=True, size=9)),
    4: (PatternFill('solid', fgColor=BLUE), Font(name='Arial', color=WHITE, bold=True, size=9)),
}

center = Alignment(horizontal='center', vertical='center')
left_al = Alignment(horizontal='left', vertical='center')
wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ===== Upper limb data =====
upper_muscles = [
    ('前鋸筋','長胸神経','-','-',[2,1,1,0,0]),
    ('棘下筋','肩甲上神経','-','上',[4,3,0,0,0]),
    ('広背筋','胸背神経','後','中',[0,1,4,1,0]),
    ('三角筋','腋窩神経','後','上',[4,1,0,0,0]),
    ('上腕二頭筋','筋皮神経','外','上',[4,3,0,0,0]),
    ('上腕三頭筋','橈骨神経','後','中',[0,1,4,3,0]),
    ('腕橈骨筋','橈骨神経','後','上',[4,1,0,0,0]),
    ('総指伸筋','橈骨神経','後','下',[0,0,3,4,0]),
    ('尺側手根伸筋','橈骨神経','後','中下',[0,1,4,2,0]),
    ('長母指伸筋','橈骨神経','後','中下',[0,0,1,4,1]),
    ('短母指伸筋','橈骨神経','後','中下',[0,0,1,4,1]),
    ('円回内筋','正中神経','外','上',[0,4,1,0,0]),
    ('浅指屈筋','正中神経','内','下',[0,0,1,3,4]),
    ('深指屈筋(示)','正中神経','内','下',[0,0,0,1,4]),
    ('長母指屈筋','正中神経','外内','上中下',[0,0,0,3,4]),
    ('方形回内筋','正中神経','外内','上中下',[0,0,0,1,4]),
    ('短母指外転筋','正中神経','内','下',[0,0,0,1,4]),
    ('深指屈筋(小)','尺骨神経','内','下',[0,0,0,4,2]),
    ('小指外転筋','尺骨神経','内','下',[0,0,0,4,4]),
    ('橈側手根伸筋','橈骨神経','内','下',[0,4,0,0,0]),
    ('尺側手根屈筋','尺骨神経','内','下',[0,0,4,3,1]),
    ('第一背側骨間筋','尺骨神経','内','下',[0,0,0,4,2]),
    ('第一掌側骨間筋','尺骨神経','内','下',[0,0,0,3,2]),
    ('母指対立筋','正中神経','内','下',[0,0,0,0,4]),
    ('母指内転筋','尺骨神経','内','下',[0,0,0,3,3]),
    ('第１第２虫様筋','正中神経','内','下',[0,0,0,4,2]),
    ('第３第４虫様筋','尺骨神経','内','下',[0,0,0,2,4]),
]

upper_levels = ['C5','C6','C7','C8','T1']
upper_nerves = ['長胸神経','肩甲上神経','胸背神経','腋窩神経','筋皮神経','橈骨神経','正中神経','尺骨神経']

# ===== Lower limb data =====
lower_muscles = [
    ('腸腰筋','大腿神経/腰神経叢',[1,1,1,0,0,0,0]),
    ('中殿筋','上殿神経',[0,0,0,1,4,1,0]),
    ('大殿筋','下殿神経',[0,0,0,0,1,2,1]),
    ('前脛骨筋','深腓骨神経',[0,0,0,2,4,0,0]),
    ('後脛骨筋','脛骨神経',[0,0,0,1,4,0,0]),
    ('長趾屈筋','脛骨神経',[0,0,0,0,4,2,0]),
    ('内転筋群','閉鎖神経',[0,2,2,2,0,0,0]),
    ('大腿四頭筋','大腿神経',[0,0,3,3,0,0,0]),
    ('大腿屈筋群','脛骨神経/総腓骨神経',[0,0,0,0,2,2,1]),
    ('足趾背屈','深腓骨神経',[0,0,0,0,4,0,0]),
    ('母趾背屈','深腓骨神経',[0,0,0,0,4,0,0]),
    ('腓骨筋群','浅腓骨神経',[0,0,0,0,4,1,0]),
    ('下腿三頭筋','脛骨神経',[0,0,0,0,2,4,0]),
    ('短趾屈筋','脛骨神経',[0,0,0,0,0,4,2]),
    ('長母趾屈筋','脛骨神経',[0,0,0,0,4,2,0]),
    ('短母趾屈筋','脛骨神経',[0,0,0,0,0,4,2]),
]

lower_levels = ['L1','L2','L3','L4','L5','S1','S2']
lower_nerves = ['大腿神経/腰神経叢','上殿神経','下殿神経','深腓骨神経','脛骨神経','閉鎖神経','大腿神経','脛骨神経/総腓骨神経','浅腓骨神経']


def build_upper_sheet(ws):
    ws.title = '上肢'
    num_muscles = len(upper_muscles)
    data_start = 5  # row
    data_end = data_start + num_muscles - 1  # row 31
    mmt_col = 'J'  # column J for MMT
    weight_cols = ['E','F','G','H','I']  # C5-T1

    # Title
    ws.merge_cells('A1:P1')
    c = ws['A1']
    c.value = '🔍 神経障害部位推測ツール — 上肢'
    c.font = title_font
    c.fill = navy_fill
    c.alignment = center

    ws.merge_cells('A2:P2')
    c = ws['A2']
    c.value = 'MMT列（黄色セル）に筋力値（0〜5）を入力すると、右側の結果セクションに自動で解析結果が表示されます'
    c.font = subtitle_font
    c.fill = PatternFill('solid', fgColor=BLUE)
    c.alignment = center

    # Headers row 4
    headers = ['筋肉','支配神経','束(Cord)','幹(Trunk)'] + upper_levels + ['MMT']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = header_font
        c.fill = navy_fill
        c.alignment = center
        c.border = thin_border
    # MMT header special
    ws.cell(row=4, column=10).fill = PatternFill('solid', fgColor='FFF5F5')
    ws.cell(row=4, column=10).font = Font(name='Arial', bold=True, color=RED, size=10)

    # Data rows
    for ri, (name, nerve, cord, trunk, weights) in enumerate(upper_muscles):
        r = data_start + ri
        ws.cell(row=r, column=1, value=name).font = Font(name='Arial', bold=True, size=10)
        ws.cell(row=r, column=1).alignment = left_al
        ws.cell(row=r, column=1).border = thin_border

        ws.cell(row=r, column=2, value=nerve).font = normal_font
        ws.cell(row=r, column=2).alignment = left_al
        ws.cell(row=r, column=2).border = thin_border

        ws.cell(row=r, column=3, value=cord).font = normal_font
        ws.cell(row=r, column=3).alignment = center
        ws.cell(row=r, column=3).border = thin_border

        ws.cell(row=r, column=4, value=trunk).font = normal_font
        ws.cell(row=r, column=4).alignment = center
        ws.cell(row=r, column=4).border = thin_border

        for wi, w in enumerate(weights):
            c = ws.cell(row=r, column=5+wi, value=w)
            fill, font = weight_fills[w]
            c.fill = fill
            c.font = font
            c.alignment = center
            c.border = thin_border

        # MMT input cell
        mc = ws.cell(row=r, column=10)
        mc.fill = input_fill
        mc.font = input_font
        mc.alignment = center
        mc.border = red_border

    # Data validation for MMT
    dv = DataValidation(type='whole', operator='between', formula1=0, formula2=5, allow_blank=True)
    dv.error = 'MMTは0〜5の整数を入力してください'
    dv.errorTitle = '入力エラー'
    ws.add_data_validation(dv)
    dv.add(f'J{data_start}:J{data_end}')

    # Column widths
    widths = {'A':18,'B':16,'C':10,'D':10,'E':5,'F':5,'G':5,'H':5,'I':5,'J':8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ===== Results Section =====
    res_col = 12  # Column L

    # Section title
    ws.merge_cells(start_row=4, start_column=res_col, end_row=4, end_column=res_col+3)
    c = ws.cell(row=4, column=res_col, value='📊 解析結果')
    c.font = Font(name='Arial', bold=True, size=14, color=WHITE)
    c.fill = navy_fill
    c.alignment = center

    # -- Spinal Level Scores --
    r = 6
    ws.cell(row=r, column=res_col, value='🔷 脊髄レベル別スコア').font = section_font
    r += 1
    for h_i, h_val in enumerate(['レベル','スコア','棒グラフ']):
        c = ws.cell(row=r, column=res_col+h_i, value=h_val)
        c.font = label_font
        c.fill = PatternFill('solid', fgColor='E8ECF1')
        c.alignment = center
    r += 1

    for li, lv in enumerate(upper_levels):
        ws.cell(row=r, column=res_col, value=lv).font = label_font
        ws.cell(row=r, column=res_col).alignment = center

        wcol = weight_cols[li]
        formula = f'=SUMPRODUCT(({mmt_col}{data_start}:{mmt_col}{data_end}<>"")*({mmt_col}{data_start}:{mmt_col}{data_end}<5)*(5-{mmt_col}{data_start}:{mmt_col}{data_end})*{wcol}{data_start}:{wcol}{data_end})'
        ws.cell(row=r, column=res_col+1, value=formula).font = Font(name='Arial', bold=True, size=11)
        ws.cell(row=r, column=res_col+1).alignment = center

        bar_formula = f'=REPT("█",MIN({get_column_letter(res_col+1)}{r},30))'
        ws.cell(row=r, column=res_col+2, value=bar_formula).font = bar_blue
        ws.cell(row=r, column=res_col+2).alignment = left_al
        r += 1

    # -- Peripheral Nerve --
    r += 1
    ws.cell(row=r, column=res_col, value='🔶 末梢神経別 障害筋数').font = section_font
    r += 1
    for h_i, h_val in enumerate(['神経名','障害筋数','棒グラフ']):
        c = ws.cell(row=r, column=res_col+h_i, value=h_val)
        c.font = label_font
        c.fill = PatternFill('solid', fgColor='E8ECF1')
        c.alignment = center
    r += 1

    for nv in upper_nerves:
        ws.cell(row=r, column=res_col, value=nv).font = normal_font
        ws.cell(row=r, column=res_col).alignment = left_al

        formula = f'=COUNTIFS(B{data_start}:B{data_end},"{nv}",{mmt_col}{data_start}:{mmt_col}{data_end},"<5",{mmt_col}{data_start}:{mmt_col}{data_end},"<>")'
        ws.cell(row=r, column=res_col+1, value=formula).font = Font(name='Arial', bold=True, size=11)
        ws.cell(row=r, column=res_col+1).alignment = center

        bar_formula = f'=REPT("█",{get_column_letter(res_col+1)}{r})'
        ws.cell(row=r, column=res_col+2, value=bar_formula).font = bar_orange
        ws.cell(row=r, column=res_col+2).alignment = left_al
        r += 1

    # -- Cord --
    r += 1
    ws.cell(row=r, column=res_col, value='🟢 神経束（Cord）別 障害筋数').font = section_font
    r += 1
    for h_i, h_val in enumerate(['神経束','障害筋数','棒グラフ']):
        c = ws.cell(row=r, column=res_col+h_i, value=h_val)
        c.font = label_font
        c.fill = PatternFill('solid', fgColor='E8ECF1')
        c.alignment = center
    r += 1

    cord_data = [('外側束','*外*'),('内側束','*内*'),('後束','*後*')]
    for label, pattern in cord_data:
        ws.cell(row=r, column=res_col, value=label).font = normal_font
        ws.cell(row=r, column=res_col).alignment = left_al

        formula = f'=COUNTIFS(C{data_start}:C{data_end},"{pattern}",{mmt_col}{data_start}:{mmt_col}{data_end},"<5",{mmt_col}{data_start}:{mmt_col}{data_end},"<>")'
        ws.cell(row=r, column=res_col+1, value=formula).font = Font(name='Arial', bold=True, size=11)
        ws.cell(row=r, column=res_col+1).alignment = center

        bar_formula = f'=REPT("█",{get_column_letter(res_col+1)}{r})'
        ws.cell(row=r, column=res_col+2, value=bar_formula).font = bar_green
        ws.cell(row=r, column=res_col+2).alignment = left_al
        r += 1

    # -- Trunk --
    r += 1
    ws.cell(row=r, column=res_col, value='🟣 神経幹（Trunk）別 障害筋数').font = section_font
    r += 1
    for h_i, h_val in enumerate(['神経幹','障害筋数','棒グラフ']):
        c = ws.cell(row=r, column=res_col+h_i, value=h_val)
        c.font = label_font
        c.fill = PatternFill('solid', fgColor='E8ECF1')
        c.alignment = center
    r += 1

    trunk_data = [('上幹','*上*'),('中幹','*中*'),('下幹','*下*')]
    for label, pattern in trunk_data:
        ws.cell(row=r, column=res_col, value=label).font = normal_font
        ws.cell(row=r, column=res_col).alignment = left_al

        formula = f'=COUNTIFS(D{data_start}:D{data_end},"{pattern}",{mmt_col}{data_start}:{mmt_col}{data_end},"<5",{mmt_col}{data_start}:{mmt_col}{data_end},"<>")'
        ws.cell(row=r, column=res_col+1, value=formula).font = Font(name='Arial', bold=True, size=11)
        ws.cell(row=r, column=res_col+1).alignment = center

        bar_formula = f'=REPT("█",{get_column_letter(res_col+1)}{r})'
        ws.cell(row=r, column=res_col+2, value=bar_formula).font = bar_purple
        ws.cell(row=r, column=res_col+2).alignment = left_al
        r += 1

    # Result column widths
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 35

    # Freeze panes
    ws.freeze_panes = 'A5'


def build_lower_sheet(ws):
    ws.title = '下肢'
    num_muscles = len(lower_muscles)
    data_start = 5
    data_end = data_start + num_muscles - 1  # row 20
    mmt_col = 'J'
    weight_cols = ['C','D','E','F','G','H','I']  # L1-S2

    # Title
    ws.merge_cells('A1:N1')
    c = ws['A1']
    c.value = '🔍 神経障害部位推測ツール — 下肢'
    c.font = title_font
    c.fill = navy_fill
    c.alignment = center

    ws.merge_cells('A2:N2')
    c = ws['A2']
    c.value = 'MMT列（黄色セル）に筋力値（0〜5）を入力すると、右側の結果セクションに自動で解析結果が表示されます'
    c.font = subtitle_font
    c.fill = PatternFill('solid', fgColor=BLUE)
    c.alignment = center

    # Headers row 4: 筋肉, 支配神経, L1-S2(7), MMT
    headers = ['筋肉','支配神経'] + lower_levels + ['MMT']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = header_font
        c.fill = navy_fill
        c.alignment = center
        c.border = thin_border
    ws.cell(row=4, column=10).fill = PatternFill('solid', fgColor='FFF5F5')
    ws.cell(row=4, column=10).font = Font(name='Arial', bold=True, color=RED, size=10)

    # Data
    for ri, (name, nerve, weights) in enumerate(lower_muscles):
        r = data_start + ri
        ws.cell(row=r, column=1, value=name).font = Font(name='Arial', bold=True, size=10)
        ws.cell(row=r, column=1).alignment = left_al
        ws.cell(row=r, column=1).border = thin_border

        ws.cell(row=r, column=2, value=nerve).font = normal_font
        ws.cell(row=r, column=2).alignment = left_al
        ws.cell(row=r, column=2).border = thin_border

        for wi, w in enumerate(weights):
            c = ws.cell(row=r, column=3+wi, value=w)
            fill, font = weight_fills[w]
            c.fill = fill
            c.font = font
            c.alignment = center
            c.border = thin_border

        mc = ws.cell(row=r, column=10)
        mc.fill = input_fill
        mc.font = input_font
        mc.alignment = center
        mc.border = red_border

    dv = DataValidation(type='whole', operator='between', formula1=0, formula2=5, allow_blank=True)
    dv.error = 'MMTは0〜5の整数を入力してください'
    dv.errorTitle = '入力エラー'
    ws.add_data_validation(dv)
    dv.add(f'J{data_start}:J{data_end}')

    widths = {'A':16,'B':22,'C':5,'D':5,'E':5,'F':5,'G':5,'H':5,'I':5,'J':8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ===== Results =====
    res_col = 12

    ws.merge_cells(start_row=4, start_column=res_col, end_row=4, end_column=res_col+3)
    c = ws.cell(row=4, column=res_col, value='📊 解析結果')
    c.font = Font(name='Arial', bold=True, size=14, color=WHITE)
    c.fill = navy_fill
    c.alignment = center

    r = 6
    ws.cell(row=r, column=res_col, value='🔷 脊髄レベル別スコア').font = section_font
    r += 1
    for h_i, h_val in enumerate(['レベル','スコア','棒グラフ']):
        c = ws.cell(row=r, column=res_col+h_i, value=h_val)
        c.font = label_font
        c.fill = PatternFill('solid', fgColor='E8ECF1')
        c.alignment = center
    r += 1

    for li, lv in enumerate(lower_levels):
        ws.cell(row=r, column=res_col, value=lv).font = label_font
        ws.cell(row=r, column=res_col).alignment = center

        wcol = weight_cols[li]
        formula = f'=SUMPRODUCT(({mmt_col}{data_start}:{mmt_col}{data_end}<>"")*({mmt_col}{data_start}:{mmt_col}{data_end}<5)*(5-{mmt_col}{data_start}:{mmt_col}{data_end})*{wcol}{data_start}:{wcol}{data_end})'
        ws.cell(row=r, column=res_col+1, value=formula).font = Font(name='Arial', bold=True, size=11)
        ws.cell(row=r, column=res_col+1).alignment = center

        bar_formula = f'=REPT("█",MIN({get_column_letter(res_col+1)}{r},30))'
        ws.cell(row=r, column=res_col+2, value=bar_formula).font = bar_blue
        ws.cell(row=r, column=res_col+2).alignment = left_al
        r += 1

    r += 1
    ws.cell(row=r, column=res_col, value='🔶 末梢神経別 障害筋数').font = section_font
    r += 1
    for h_i, h_val in enumerate(['神経名','障害筋数','棒グラフ']):
        c = ws.cell(row=r, column=res_col+h_i, value=h_val)
        c.font = label_font
        c.fill = PatternFill('solid', fgColor='E8ECF1')
        c.alignment = center
    r += 1

    for nv in lower_nerves:
        ws.cell(row=r, column=res_col, value=nv).font = normal_font
        ws.cell(row=r, column=res_col).alignment = left_al

        formula = f'=COUNTIFS(B{data_start}:B{data_end},"{nv}",{mmt_col}{data_start}:{mmt_col}{data_end},"<5",{mmt_col}{data_start}:{mmt_col}{data_end},"<>")'
        ws.cell(row=r, column=res_col+1, value=formula).font = Font(name='Arial', bold=True, size=11)
        ws.cell(row=r, column=res_col+1).alignment = center

        bar_formula = f'=REPT("█",{get_column_letter(res_col+1)}{r})'
        ws.cell(row=r, column=res_col+2, value=bar_formula).font = bar_orange
        ws.cell(row=r, column=res_col+2).alignment = left_al
        r += 1

    ws.column_dimensions['L'].width = 24
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 35

    ws.freeze_panes = 'A5'


def build_instruction_sheet(ws):
    ws.title = '使い方'

    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = '📖 神経障害部位推測ツール — 使い方'
    c.font = Font(name='Arial', bold=True, size=16, color=NAVY)
    c.alignment = left_al

    instructions = [
        ('', ''),
        ('使い方', ''),
        ('1.', '「上肢」または「下肢」シートを開く'),
        ('2.', 'MMT列（黄色のセル）に筋力値（0〜5）を入力'),
        ('3.', '右側の結果セクションに自動で解析結果が表示される'),
        ('4.', '脊髄レベル、末梢神経、神経束、神経幹の観点から局在を推定'),
        ('5.', '棒グラフ（█）でスコアの大きさを視覚的に確認'),
        ('', ''),
        ('⚠️ 注意', '本ツールはあくまで補助的な推測です。臨床所見・画像検査・電気生理学的検査と合わせて総合的に判断してください。'),
        ('', ''),
        ('重み付けの意味', ''),
        ('0', 'その脊髄レベルからの支配がない'),
        ('1', '支配の可能性が低い（副次的）'),
        ('2', '中程度の支配'),
        ('3', '主要な支配の一つ'),
        ('4', '最も強い支配（主要な脊髄レベル）'),
        ('', ''),
        ('束（Cord）の略称', ''),
        ('外', '外側束（Lateral cord）'),
        ('内', '内側束（Medial cord）'),
        ('後', '後束（Posterior cord）'),
        ('外内', '外側束＋内側束の両方から支配'),
        ('', ''),
        ('幹（Trunk）の略称', ''),
        ('上', '上幹（Upper trunk）— C5, C6'),
        ('中', '中幹（Middle trunk）— C7'),
        ('下', '下幹（Lower trunk）— C8, T1'),
        ('上中下', '全ての幹から支配'),
        ('中下', '中幹＋下幹から支配'),
    ]

    for ri, (col_a, col_b) in enumerate(instructions, 3):
        ca = ws.cell(row=ri, column=1, value=col_a)
        cb = ws.cell(row=ri, column=2, value=col_b)
        if col_a in ('使い方','重み付けの意味','束（Cord）の略称','幹（Trunk）の略称','⚠️ 注意'):
            ca.font = Font(name='Arial', bold=True, size=12, color=NAVY)
        else:
            ca.font = Font(name='Arial', bold=True, size=10)
        cb.font = normal_font
        cb.alignment = wrap

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60


# ===== Build workbook =====
build_upper_sheet(wb.active)
build_lower_sheet(wb.create_sheet())
build_instruction_sheet(wb.create_sheet())

output = r'C:\Users\jsber\OneDrive\Documents\Claude_task\神経障害部位推測ツール.xlsx'
wb.save(output)
print(f'Saved: {output}')
